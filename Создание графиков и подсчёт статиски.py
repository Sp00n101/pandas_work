import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from openpyxl import load_workbook
from math import sqrt, pi, exp


def text(zna, poten):

    text = 'Process Data   \n'
    for i in range(len(zna)):
        z = list(zna.items())[i]
        text += z[0] + (7 -len(z[0]))*" " + " "*(8 - len(str(round(z[1], 5)))) + str(round(z[1], 5)) + "\n"
    text += '\n\nCapability \n'
    for i in range(len(poten)):
        p = list(poten.items())[i]
        text += p[0] + " " * ((3 - len(p[0]))) + " "*(8 - len(str(round(p[1], 2))))  + str(round(p[1], 2)) + "\n"
    print(text)
    return text


def normrasp(X,q,u,mnog):
    y_data = []
    for i in X:
        y = mnog*1/q/sqrt(2*pi)*exp(-(i-u)*(i-u)/2/q/q)
        y_data.append(y)
    print(len(X),len(y_data))
    return y_data
def num_after_point(x):
    s = str(x)
    if not '.' in s:
        return 0
    return len(s) - s.index('.') - 1


def bins_range(x, d, c=1):
    bins_ran = []
    for i in np.arange(round((min(x)) - d,num_after_point(d)), round((max(x)) + d,num_after_point(d)),d/c):
        bins_ran.append(i + d/2/c)
    return bins_ran


def znach(x, c = None):
    if c != None:
        usl = USL1
        lsl = LSL1
        target = Target1
    else:
        usl = USL
        lsl = LSL
        target = Target
    mean = x.mean()
    disp = 0
    mr = 0
    for j in range(1,len(x)):
        mr += abs(df.loc[j,"AVG"] - df.loc[j-1,"AVG"])
    for i in x:
        disp += (i - mean) ** 2
    StO = sqrt(disp / (len(x) -1))
    return { "LSL": lsl, "Target": target, "USL": usl, "Mean": mean, "N": len(x), "StDev": StO,}


def potential(c):
    n = c['N']
    tar = c['Target']
    so = c['StDev']
    lsl = c['LSL']
    usl = c['USL']
    mean = c["Mean"]
    pp = (USL-LSL)/6/so
    ppl = (mean-lsl)/3/so
    ppu = (usl-mean)/3/so
    ppk = min(ppl, ppu)
    cpm = (USL-LSL)/6/sqrt(so*so+(mean-tar)*(mean-tar))
    print((USL-LSL))
    return {'Pp': pp, 'PPL': ppl, 'PPU': ppu, 'PPk': ppk, 'Cpm': cpm}

def gist(DF,c = None):
    if c != None:
        usl = USL1
        lsl = LSL1
        target = Target1
        bin = bins_range(DF, ws.cell(2, 9).value, 2)
    else:
        usl = USL
        lsl = LSL
        target = Target
        bin = bins_range(DF, ws.cell(2, 9).value)
    Q = znach(DF,c)["StDev"]
    U = znach(DF,c)["Mean"]
    x_axis = np.arange(U - max(U - lsl, usl - U), U + max(U - lsl, usl - U), 0.001)
    plt.figure(figsize=(9, 5))
    plt.hist(DF, bins=bin, edgecolor='blue')  # bins - количество столбцов в гистограмме
    plt.xlabel('Значение')
    plt.ylabel('Частота')
    plt.title("Гистограмма")
    plt.axvline(x=usl, linestyle='dashed', color='red', linewidth=1, label = "Границы LSL-USL")
    plt.axvline(x=lsl,linestyle='dashed', color='red', linewidth=1)
    plt.axvline(x=target, linestyle='dashed', color='green', linewidth=2, label = "Target")
    plt.plot(x_axis, normrasp(x_axis, Q, U,len(DF)), color = "red")
    #plt. text(12.1, 1, text(znach(DF),potential(znach(DF))), fontsize=9)
    plt.text(.99, .99, text(znach(DF,c),potential(znach(DF,c))), ha='right', va='top', transform=plt.gca().transAxes, fontstyle='normal', family='monospace')
    plt.show()


files_in_directory = os.listdir("./")
file_name = ""
sheet_name = "Данные"

for file in files_in_directory:
    if file == "Данные_для_обработки.xlsx":
        file_name = "./" + file
        break

if file_name != '':
    wb = load_workbook(file_name, read_only=True)
    ws = wb[sheet_name]

    LSL = ws.cell(2, 5).value
    USL = ws.cell(2, 6).value
    Target = ws.cell(2, 7).value

    LSL1 = ws.cell(3, 5).value
    USL1 = ws.cell(3, 6).value
    Target1 = ws.cell(3, 7).value


    df = pd.read_excel(file_name, sheet_name=sheet_name, skiprows=0, usecols="A:B")
    df = df.dropna(axis=0, how='all')
    # Удаление столбцов, в которых отсутствуют данные
    df = df.dropna(axis=1, how='all')
    df.columns = ['X', 'Y']
    df['AVG'] = df[['X', 'Y']].mean(axis=1)
    df['Diff'] = pd.Series( df['X'] - df['Y'])

    print(potential(znach(df.AVG)))
    gist(df.AVG)
    gist(df.Diff, 0)






